from flask import Flask, render_template, request, send_file, redirect, url_for, flash
from werkzeug.utils import secure_filename
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import uuid

app = Flask(__name__)
app.secret_key = "secret123"

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_FOLDER = BASE_DIR / "uploads"
OUTPUT_FOLDER = BASE_DIR / "outputs"

UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {"xlsx", "xls", "csv"}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def clean_order_id(val):
    if pd.isna(val):
        return ""
    val = str(val).strip()
    if val.endswith(".0"):
        val = val[:-2]
    return val


def clean_phone(val):
    if pd.isna(val):
        return None

    val = str(val).strip()

    if not val or val.lower() == "nan":
        return None

    if "E+" in val or "e+" in val:
        try:
            val = str(int(float(val)))
        except Exception:
            pass

    if val.endswith(".0"):
        val = val[:-2]

    return val


def read_shopify_file(shopify_path):
    if shopify_path.suffix.lower() == ".csv":
        shopify = pd.read_csv(shopify_path, dtype=str)
    else:
        shopify = pd.read_excel(shopify_path, dtype=str)

    shopify.columns = shopify.columns.str.strip()

    required_cols = ["Name", "Email", "Phone"]
    missing = [c for c in required_cols if c not in shopify.columns]
    if missing:
        raise Exception(f"Shopify file is missing columns: {', '.join(missing)}")

    shopify["Name"] = shopify["Name"].apply(clean_order_id)
    shopify["Email"] = shopify["Email"].astype(str).str.strip()
    shopify["Phone"] = shopify["Phone"].apply(clean_phone)

    shopify["Email"] = shopify["Email"].replace({"": None, "nan": None, "None": None})
    shopify["Phone"] = shopify["Phone"].replace({"": None, "nan": None, "None": None})

    shopify = shopify[shopify["Name"] != ""].copy()
    shopify = shopify.drop_duplicates(subset=["Name"], keep="last")

    return shopify


def find_matching_column(headers, possible_names):
    for header_text, col_num in headers.items():
        for name in possible_names:
            if header_text.lower().strip() == name.lower().strip():
                return col_num
    return None


def save_changed_ids_report(changed_ids, prefix):
    if not changed_ids:
        return None

    report_name = f"{prefix}_{uuid.uuid4().hex}.xlsx"
    report_path = OUTPUT_FOLDER / report_name

    pd.DataFrame({"Order ID": changed_ids}).to_excel(report_path, index=False)
    return report_name


def update_master(shopify_path, master_path, sheet_name, fill_blanks_only=True):
    shopify = read_shopify_file(shopify_path)
    shopify_dict = shopify.set_index("Name")[["Email", "Phone"]].to_dict("index")

    wb = load_workbook(master_path)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise Exception(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]

    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value is not None:
            headers[str(cell_value).strip()] = col

    order_col = find_matching_column(headers, ["Order ID", "OrderId", "Order Id"])
    email_col = find_matching_column(headers, ["Email", "Email Address"])
    phone_col = find_matching_column(headers, ["Mobile no", "Mobile No", "Mobile", "Phone", "Phone Number"])

    if not order_col or not email_col or not phone_col:
        wb.close()
        raise Exception("Master sheet is missing required columns: Order ID, Email, Mobile no")

    matched = 0
    unmatched = 0
    email_updated = 0
    mobile_updated = 0

    changed_email_ids = []
    changed_mobile_ids = []
    unmatched_order_ids = []

    for row in range(2, ws.max_row + 1):
        order_id = clean_order_id(ws.cell(row=row, column=order_col).value)

        if not order_id:
            continue

        if order_id not in shopify_dict:
            unmatched += 1
            unmatched_order_ids.append(order_id)
            continue

        matched += 1
        record = shopify_dict[order_id]

        new_email = record.get("Email")
        new_phone = record.get("Phone")

        email_cell = ws.cell(row=row, column=email_col)
        phone_cell = ws.cell(row=row, column=phone_col)

        current_email = email_cell.value
        current_phone = phone_cell.value

        current_email_clean = "" if current_email is None else str(current_email).strip()
        new_email_clean = "" if new_email is None else str(new_email).strip()

        current_phone_clean = clean_phone(current_phone)
        new_phone_clean = clean_phone(new_phone)

        # EMAIL UPDATE
        if new_email_clean:
            if fill_blanks_only:
                if current_email is None or current_email_clean == "":
                    email_cell.value = new_email_clean
                    email_updated += 1
                    changed_email_ids.append(order_id)
            else:
                if current_email_clean != new_email_clean:
                    email_cell.value = new_email_clean
                    email_updated += 1
                    changed_email_ids.append(order_id)

        # MOBILE UPDATE
        if new_phone_clean:
            if fill_blanks_only:
                if current_phone is None or str(current_phone).strip() == "":
                    phone_cell.value = new_phone_clean
                    mobile_updated += 1
                    changed_mobile_ids.append(order_id)
            else:
                if current_phone_clean != new_phone_clean:
                    phone_cell.value = new_phone_clean
                    mobile_updated += 1
                    changed_mobile_ids.append(order_id)

    output_name = f"updated_{uuid.uuid4().hex}.xlsx"
    output_path = OUTPUT_FOLDER / output_name

    wb.template = False
    wb.save(str(output_path))
    wb.close()

    changed_mobile_file = save_changed_ids_report(changed_mobile_ids, "changed_mobile_ids")
    changed_email_file = save_changed_ids_report(changed_email_ids, "changed_email_ids")
    unmatched_file = save_changed_ids_report(unmatched_order_ids, "unmatched_order_ids")

    return output_name, {
        "matched": matched,
        "unmatched": unmatched,
        "email_updated": email_updated,
        "mobile_updated": mobile_updated,
        "changed_mobile_ids": changed_mobile_ids,
        "changed_email_ids": changed_email_ids,
        "unmatched_order_ids": unmatched_order_ids,
        "changed_mobile_file": changed_mobile_file,
        "changed_email_file": changed_email_file,
        "unmatched_file": unmatched_file,
    }


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        step = request.form.get("step", "upload")

        if step == "upload":
            shopify_file = request.files.get("shopify_file")
            master_file = request.files.get("master_file")
            fill_blanks_only = request.form.get("fill_blanks_only") == "on"

            if not shopify_file or shopify_file.filename == "":
                flash("Please upload Shopify file", "error")
                return redirect(url_for("index"))

            if not master_file or master_file.filename == "":
                flash("Please upload Master file", "error")
                return redirect(url_for("index"))

            if not allowed_file(shopify_file.filename):
                flash("Invalid Shopify file type", "error")
                return redirect(url_for("index"))

            if not allowed_file(master_file.filename):
                flash("Invalid Master file type", "error")
                return redirect(url_for("index"))

            shopify_saved_name = f"{uuid.uuid4().hex}_{secure_filename(shopify_file.filename)}"
            master_saved_name = f"{uuid.uuid4().hex}_{secure_filename(master_file.filename)}"

            shopify_path = UPLOAD_FOLDER / shopify_saved_name
            master_path = UPLOAD_FOLDER / master_saved_name

            shopify_file.save(shopify_path)
            master_file.save(master_path)

            wb = load_workbook(master_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()

            return render_template(
                "index.html",
                sheets=sheets,
                shopify_saved_name=shopify_saved_name,
                master_saved_name=master_saved_name,
                fill_blanks_only=fill_blanks_only,
            )

        elif step == "process":
            shopify_saved_name = request.form.get("shopify_saved_name", "").strip()
            master_saved_name = request.form.get("master_saved_name", "").strip()
            sheet_name = request.form.get("sheet_name", "").strip()
            fill_blanks_only = request.form.get("fill_blanks_only") == "on"

            if not shopify_saved_name or not master_saved_name:
                flash("Uploaded files not found. Please upload again.", "error")
                return redirect(url_for("index"))

            if not sheet_name:
                flash("Please select a sheet", "error")
                return redirect(url_for("index"))

            shopify_path = UPLOAD_FOLDER / shopify_saved_name
            master_path = UPLOAD_FOLDER / master_saved_name

            if not shopify_path.exists() or not master_path.exists():
                flash("Uploaded files are missing. Please upload again.", "error")
                return redirect(url_for("index"))

            try:
                output_name, result = update_master(
                    shopify_path=shopify_path,
                    master_path=master_path,
                    sheet_name=sheet_name,
                    fill_blanks_only=fill_blanks_only,
                )

                return render_template(
                    "index.html",
                    success=True,
                    result=result,
                    file=output_name,
                )
            except Exception as e:
                flash(str(e), "error")
                return redirect(url_for("index"))

    return render_template("index.html")


@app.route("/download/<filename>")
def download(filename):
    path = OUTPUT_FOLDER / filename
    if not path.exists():
        flash("Output file not found", "error")
        return redirect(url_for("index"))
    return send_file(path, as_attachment=True, download_name=path.name)


if __name__ == "__main__":
    app.run(debug=True)